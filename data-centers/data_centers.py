#!/usr/bin/env python
# coding: utf-8

# In[4]:


# ==== Datacenters.com (Australia) scraper ====
# Fast: Selenium only for 7 listing pages (images blocked, eager load)
# Fast: Details + images fetched in parallel over HTTP
# Map Link: computed "place" link so Google Maps opens with a pin
# Saves to Excel AFTER EACH datacenter processed
# Missing fields -> empty strings

from pathlib import Path
import time, re, traceback, threading
from urllib.parse import urljoin, urlsplit, quote_plus
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
import pandas as pd

# Excel append helper
from openpyxl import Workbook, load_workbook

# Selenium only for listing pagination (7 pages)
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ------------------ CONFIG ------------------
BASE_LISTING_URL = "https://www.datacenters.com/locations/australia"
BASE_DOMAIN      = "https://www.datacenters.com"

BASE_DIR   = Path(r"D:\Upwork\Project 6 (Australia)")
LOGO_DIR   = BASE_DIR / "images" / "logo"
MEDIA_DIR  = BASE_DIR / "images" / "media"
EXCEL_PATH = BASE_DIR / "Australia Datacenters.xlsx"

MAX_PAGES            = 7     # stop strictly after page 7
DETAIL_MAX_WORKERS   = 18    # tune 12-24 for your network/CPU
REQUESTS_TIMEOUT_S   = 25
HEADLESS             = True  # set False if you want to watch

BASE_DIR.mkdir(parents=True, exist_ok=True)
LOGO_DIR.mkdir(parents=True, exist_ok=True)
MEDIA_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}


# ------------------ Utility helpers ------------------
def slugify(value: str, allow_dot=False) -> str:
    value = (value or "").strip()
    value = re.sub(r"[^\w\-. ]" if allow_dot else r"[^\w\- ]", "", value)
    value = re.sub(r"\s+", "_", value)
    return value[:160] or "file"

def filename_from_url(url: str) -> str:
    path = urlsplit(url).path
    return Path(path).name or "image"

def build_gmaps_place_link(address: str) -> str:
    """Return a Google Maps link that opens with a *pin* by searching the address."""
    if not address:
        return ""
    return f"https://www.google.com/maps/place/{quote_plus(address)}"

# Thread-local pooled HTTP sessions for speed
_thread_local = threading.local()
def get_session() -> requests.Session:
    if not hasattr(_thread_local, "session"):
        s = requests.Session()
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry
        retries = Retry(
            total=2, backoff_factor=0.25,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=frozenset(["GET", "HEAD"])
        )
        adapter = HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=retries)
        s.mount("https://", adapter)
        s.mount("http://", adapter)
        s.headers.update(HEADERS)
        _thread_local.session = s
    return _thread_local.session

def save_image(url: str, folder: Path, name_hint: str, session: requests.Session = None) -> str:
    """Download image and return just the filename; '' on failure/missing."""
    if not url:
        return ""
    try:
        orig = filename_from_url(url)
        ext = Path(orig).suffix or ".png"
        out = folder / f"{slugify(name_hint, allow_dot=True)}{ext}"
        if out.exists():
            base = out.stem
            for i in range(2, 10000):
                cand = folder / f"{base}_{i}{ext}"
                if not cand.exists():
                    out = cand
                    break
        sess = session or get_session()
        with sess.get(url, timeout=REQUESTS_TIMEOUT_S, stream=True) as r:
            r.raise_for_status()
            with open(out, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
        return out.name
    except Exception as e:
        print(f"[warn] image download failed: {url} -> {e}")
        return ""


# ------------------ Excel incremental writer ------------------
class ExcelAppender:
    def __init__(self, path: Path, headers: list[str]):
        self.path = Path(path)
        self.headers = headers
        if self.path.exists():
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
            # If header missing (corrupted), add it
            if self.ws.max_row == 0:
                self.ws.append(self.headers)
                self.wb.save(self.path)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(self.headers)
            self.wb.save(self.path)

    def append(self, row_dict: dict):
        row = [row_dict.get(h, "") for h in self.headers]
        self.ws.append(row)
        self.wb.save(self.path)

    def close(self):
        try:
            self.wb.save(self.path)
            self.wb.close()
        except:
            pass


# ------------------ Listing (Selenium, super light) ------------------
def make_driver(headless=True):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    # Faster/eager loads + block images
    try:
        opts.set_capability("pageLoadStrategy", "eager")
    except Exception:
        pass
    prefs = {"profile.managed_default_content_settings.images": 2}
    opts.add_experimental_option("prefs", prefs)

    # Hardening
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-features=site-per-process")
    opts.add_argument(f"user-agent={HEADERS['User-Agent']}")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    driver.set_page_load_timeout(30)
    driver.implicitly_wait(5)
    return driver

def wait_for_cards(driver):
    sel = "a[href^='/'][class*='flex'][class*='rounded'][class*='border'][class*='p-2']"
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
    return sel

def parse_listing_cards_from_html(html: str):
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select("a[href^='/'][class*='flex'][class*='rounded'][class*='border'][class*='p-2']")
    rows = []
    for a in cards:
        href = (a.get("href") or "").strip()
        full_url = urljoin(BASE_DOMAIN, href) if href.startswith("/") else href or ""

        facility_el = a.select_one("div.text.font-medium")
        facility = facility_el.get_text(strip=True) if facility_el else ""

        smalls = a.select("div.text-xs.text-gray-500")
        operator = smalls[0].get_text(strip=True) if smalls else ""
        address  = smalls[-1].get_text(strip=True) if smalls else ""

        logo_el  = a.select_one("div.relative div.absolute img")
        logo_url = (logo_el.get("src") or "").strip() if logo_el else ""

        media_url = ""
        rel = a.select_one("div.relative")
        if rel:
            pics = rel.select("picture img")
            if pics:
                media_url = (pics[-1].get("src") or "").strip()

        rows.append({
            "Operator": operator or "",
            "Facility": facility or "",
            "Address": address or "",
            "Data Center URL": href or "",
            "Full URL": full_url or "",
            "Logo Image URL": logo_url or "",
            "Media Image URL": media_url or ""
        })
    return rows

def click_next_and_wait(driver, listing_selector) -> bool:
    old = {el.get_attribute("href") for el in driver.find_elements(By.CSS_SELECTOR, listing_selector)}
    btns = driver.find_elements(By.XPATH, "//button[.//div[normalize-space()='➞']]")
    btns = [b for b in btns if b.is_displayed()]
    if not btns:
        return False
    next_btn = btns[-1]
    if (next_btn.get_attribute("disabled") or "").lower() in ("true", "disabled"):
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_btn)
        time.sleep(0.15)
        next_btn.click()
    except Exception:
        try:
            driver.execute_script("arguments[0].click();", next_btn)
        except Exception:
            return False
    # wait until at least one new href appears
    try:
        WebDriverWait(driver, 10).until(
            lambda d: any(e.get_attribute("href") not in old
                          for e in d.find_elements(By.CSS_SELECTOR, listing_selector))
        )
        return True
    except Exception:
        return False


# ------------------ Detail page (HTTP only, fast) ------------------
def parse_detail_html(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    out = {
        "Short Details": "",
        "Long Details": "",
        "Google Map Link (page)": "",   # raw link if page provides one
        "Detail Address": "",
        "Total Space": "",
        "Colocation Space": "",
        "Total Power": "",
        "Power Density": "",
        "Nearest Airport": ""
    }

    el = soup.select_one("#contentSummary")
    out["Short Details"] = el.get_text(" ", strip=True) if el else ""

    desc = soup.select_one("#contentDescription")
    if desc:
        ps = [p.get_text(" ", strip=True) for p in desc.select("p")]
        out["Long Details"] = " ".join(ps) if ps else desc.get_text(" ", strip=True)

    # Sometimes the page already contains a google maps link
    g = soup.select_one("a[href*='maps.google.com']")
    if g:
        out["Google Map Link (page)"] = (g.get("href") or "").strip()

    # Address under the map tab block (often server-rendered even if map requires click)
    addr_block = soup.find("div", class_=lambda v: v and "LocationShowContentMapTab__addressContainer" in v)
    if addr_block:
        divs = addr_block.find_all("div")
        if divs:
            out["Detail Address"] = divs[-1].get_text(" ", strip=True)

    def stat(stat_id):
        s = soup.select_one(f"#{stat_id} strong")
        return s.get_text(" ", strip=True) if s else ""

    out["Total Space"]      = stat("totalSpace")
    out["Colocation Space"] = stat("colocationSpace")
    out["Total Power"]      = stat("power")
    out["Power Density"]    = stat("powerDensity")
    out["Nearest Airport"]  = stat("airport")

    return out

def fetch_detail_and_images(row: dict) -> dict:
    """HTTP worker: fetch detail, compute map link, download images, return final record."""
    session = get_session()
    full_url = row.get("Full URL") or urljoin(BASE_DOMAIN, row.get("Data Center URL",""))
    details = {
        "Short Details": "", "Long Details": "",
        "Google Map Link (page)": "", "Detail Address": "",
        "Total Space": "", "Colocation Space": "",
        "Total Power": "", "Power Density": "", "Nearest Airport": ""
    }
    try:
        r = session.get(full_url, timeout=REQUESTS_TIMEOUT_S)
        r.raise_for_status()
        details = parse_detail_html(r.text)
    except Exception as e:
        print(f"[warn] detail GET failed: {full_url} -> {e}")

    # Choose address to build a pin link
    addr_for_map = details.get("Detail Address") or row.get("Address") or ""
    gmaps_pin = build_gmaps_place_link(addr_for_map) if addr_for_map else ""

    # Download images
    op, fac = (row.get("Operator") or ""), (row.get("Facility") or "")
    hint = slugify(f"{op}_{fac}")
    logo_name  = save_image(row.get("Logo Image URL",""),  LOGO_DIR,  f"logo_{hint}",  session=session) if row.get("Logo Image URL") else ""
    media_name = save_image(row.get("Media Image URL",""), MEDIA_DIR, f"media_{hint}", session=session) if row.get("Media Image URL") else ""

    out = {
        "Operator": row.get("Operator","") or "",
        "Facility": row.get("Facility","") or "",
        "Address": row.get("Address","") or "",
        "Detail Address": details.get("Detail Address","") or "",
        "Data Center URL": row.get("Data Center URL","") or "",
        "Full URL": full_url or "",
        "Logo Image URL": row.get("Logo Image URL","") or "",
        "Media Image URL": row.get("Media Image URL","") or "",
        "Logo Image Name": logo_name,
        "Media Image Name": media_name,
        "Short Details": details.get("Short Details","") or "",
        "Long Details": details.get("Long Details","") or "",
        # Prefer pinned link we compute; if empty, fall back to page link if present
        "Google Map Link": gmaps_pin or details.get("Google Map Link (page)","") or "",
        "Total Space": details.get("Total Space","") or "",
        "Colocation Space": details.get("Colocation Space","") or "",
        "Total Power": details.get("Total Power","") or "",
        "Power Density": details.get("Power Density","") or "",
        "Nearest Airport": details.get("Nearest Airport","") or "",
    }
    return out


# ------------------ Run: collect listings (7 pages) ------------------
driver = make_driver(HEADLESS)
records_written = 0

# Final columns (Excel header)
cols_order = [
    "Operator", "Facility", "Address", "Detail Address",
    "Data Center URL", "Full URL",
    "Logo Image URL", "Media Image URL",
    "Logo Image Name", "Media Image Name",
    "Short Details", "Long Details", "Google Map Link",
    "Total Space", "Colocation Space", "Total Power", "Power Density", "Nearest Airport"
]
excel = ExcelAppender(EXCEL_PATH, cols_order)

all_rows = []
try:
    driver.get(BASE_LISTING_URL)
    listing_selector = wait_for_cards(driver)

    for page_ix in range(1, MAX_PAGES+1):
        print(f"Listing page {page_ix} ...")
        html = driver.page_source
        rows = parse_listing_cards_from_html(html)
        print(f"  cards: {len(rows)}")
        all_rows.extend(rows)

        if page_ix == MAX_PAGES:
            break
        if not click_next_and_wait(driver, listing_selector):
            print("  next page unavailable -> stop.")
            break
        time.sleep(0.4)
finally:
    try: driver.quit()
    except: pass

# De-duplicate by Full URL (preserve order)
seen = set()
unique_rows = []
for r in all_rows:
    u = r.get("Full URL") or urljoin(BASE_DOMAIN, r.get("Data Center URL",""))
    if u and u not in seen:
        seen.add(u)
        r["Full URL"] = u
        unique_rows.append(r)

print(f"Unique listings: {len(unique_rows)}")


# ------------------ Parallel details + incremental Excel save ------------------
with ThreadPoolExecutor(max_workers=DETAIL_MAX_WORKERS) as ex:
    futs = [ex.submit(fetch_detail_and_images, row) for row in unique_rows]
    for i, fut in enumerate(as_completed(futs), 1):
        try:
            rec = fut.result()
        except Exception as e:
            print(f"[warn] worker failed: {e}")
            continue

        # Append to Excel immediately
        excel.append(rec)
        records_written += 1
        if records_written % 5 == 0 or records_written <= 3:
            print(f"  saved {records_written}/{len(unique_rows)} rows to Excel")

excel.close()
print(f"\nDONE. Total rows written: {records_written}")
print(f"Excel -> {EXCEL_PATH}")
print(f"Logos -> {LOGO_DIR}")
print(f"Media -> {MEDIA_DIR}")


# In[ ]:





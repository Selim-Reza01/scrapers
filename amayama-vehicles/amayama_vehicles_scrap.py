#!/usr/bin/env python
# coding: utf-8

# In[14]:


import os
import time
import openpyxl
from openpyxl import Workbook
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By

# Excel output path
excel_path = "D:\\Car x Parts\\Amayama\\subaru.xlsx"

# Create Excel file if it doesn’t exist
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Make", "Model", "Chassis", "chassis-URL"])
    wb.save(excel_path)

# Load workbook
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Constants
BASE_URL = "https://www.amayama.com"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/115.0.0.0 Safari/537.36"
)

# Chrome options setup
options = uc.ChromeOptions()
options.add_argument(f'user-agent={USER_AGENT}')
options.add_argument("--disable-blink-features=AutomationControlled")

# Start browser
driver = uc.Chrome(options=options)
driver.implicitly_wait(10)

# Captcha check
def check_for_captcha():
    text = driver.page_source.lower()
    if any(t in text for t in ["verify you are human", "just a moment", "security check"]):
        print("Captcha detected! Please solve it manually.")
        input("Press Enter after solving...")

# Step 1: Open Skoda model catalog
driver.get(BASE_URL + "/en/genuine-catalogs/subaru")
check_for_captcha()
time.sleep(2)

# Step 2: Grab all model names and hrefs before navigating
model_elements = driver.find_elements(By.CSS_SELECTOR, ".list-group > a")
models = []
for model in model_elements:
    try:
        model_name = model.text.strip()
        model_href = model.get_attribute("href")
        models.append((model_name, model_href))
    except Exception as e:
        print(f"Failed to capture model: {e}")
        continue

print("Model Found:")
for name, _ in models:
    print(f"  - {name}")

# Step 3: Visit each model and extract chassis
for model_name, model_href in models:
    print(f"Processing . . . {model_name}")
    try:
        driver.get(model_href)
        check_for_captcha()
        time.sleep(2)

        epc_blocks = driver.find_elements(By.CSS_SELECTOR, ".epcModel")

        for block in epc_blocks:
            try:
                links = block.find_elements(By.CSS_SELECTOR, ".epcModel__frames a")
                for link in links:
                    chassis_name = link.text.strip()
                    chassis_url = link.get_attribute("href")
                    ws.append(["Subaru", model_name, chassis_name, chassis_url])
            except Exception as inner_err:
                print(f"Error scraping chassis from {model_name}: {inner_err}")
                continue
    except Exception as outer_err:
        print(f"Error accessing model page {model_name}: {outer_err}")
        continue

# Save results to Excel
wb.save(excel_path)
print("Scraping Complete")

# Clean up
driver.quit()


# In[ ]:





# In[16]:


import os
import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By

# File paths
input_path = "D:\\Car x Parts\\Amayama\\input.xlsx"
output_path = "D:\\Car x Parts\\Amayama\\all_output1.xlsx"

# Read input Excel
df = pd.read_excel(input_path)

# Setup browser options
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/115.0.0.0 Safari/537.36"
)

options = uc.ChromeOptions()
options.add_argument(f'user-agent={USER_AGENT}')
options.add_argument("--disable-blink-features=AutomationControlled")
# options.add_argument("--headless")  # Optional

# Start driver
driver = uc.Chrome(options=options)
driver.implicitly_wait(10)

# Captcha check
def check_for_captcha():
    text = driver.page_source.lower()
    if any(t in text for t in ["verify you are human", "just a moment", "security check"]):
        print("Captcha detected! Please solve it manually.")
        input("Press Enter after solving...")

# Prepare Excel workbook
if not os.path.exists(output_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Make", "Model", "Chassis", "chassis-URL",
        "New_Model", "Engine Model", "Engine_capacity",
        "Prod period", "Body", "Options", "URL"
    ])
    wb.save(output_path)

wb = openpyxl.load_workbook(output_path)
ws = wb.active

# Iterate through each row in input
total = len(df)
for idx, row in df.iterrows():
    print(f"Working {idx+1} out of {total}")
    print("Processing ----")

    make = str(row["Make"])
    model = str(row["Model"])
    chassis = str(row["Chassis"])
    chassis_url = str(row["chassis-URL"])

    try:
        driver.get(chassis_url)
        check_for_captcha()
        time.sleep(2)

        rows = driver.find_elements(By.CSS_SELECTOR, "tr.epcVariations__row")

        for r in rows:
            try:
                cols = r.find_elements(By.TAG_NAME, "td")
                if not cols:
                    continue

                # Defaults
                new_model = engine_model = engine_capacity = prod_period = body = options = url = "null"

                if len(cols) > 0:
                    new_model = cols[0].text.strip()
                    try:
                        url = cols[0].find_element(By.TAG_NAME, "a").get_attribute("href")
                    except:
                        url = "null"

                if len(cols) > 1:
                    engine_model = cols[1].text.strip()
                    try:
                        engine_capacity = cols[1].find_element(By.CLASS_NAME, "info-hint-new").get_attribute("data-content")
                    except:
                        engine_capacity = "null"

                if len(cols) > 2:
                    prod_period = cols[2].text.strip()

                if len(cols) > 3:
                    body = cols[3].text.strip()

                if len(cols) > 6:
                    options = cols[6].text.strip()

                # Write to Excel
                ws.append([
                    make, model, chassis, chassis_url,
                    new_model, engine_model, engine_capacity,
                    prod_period, body, options, url
                ])

            except Exception as sub_e:
                print(f"Error parsing row: {sub_e}")
                continue

    except Exception as e:
        print(f"Failed for {chassis_url}: {e}")
        continue

# Save and quit
wb.save(output_path)
driver.quit()
print("Scraping Complete ✅")


# In[ ]:





# In[ ]:





# In[4]:


import os
import time
import random
import pandas as pd
import openpyxl
from openpyxl import Workbook
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By

# ========== SETTINGS ==========
input_path = "D:\\Car x Parts\\Amayama\\new_input.xlsx"
output_base = "D:\\Car x Parts\\Amayama\\all_final_output1x"

MAX_ROWS = 1000000  # Safe buffer under Excel max rows
SLEEP_MIN, SLEEP_MAX = 0.5, 1.2  # Random sleep range
# ==============================

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/115.0.0.0 Safari/537.36"
)

options = uc.ChromeOptions()
options.add_argument(f'user-agent={USER_AGENT}')
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--disable-infobars")
options.add_argument("--disable-extensions")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
# options.add_argument("--headless")  # Optional: Uncomment to speed up with headless

driver = uc.Chrome(options=options)
driver.implicitly_wait(10)

def check_for_captcha():
    current_url = driver.current_url.lower()
    page_text = driver.page_source.lower()
    
    # Case 1: Specific CAPTCHA URL
    if "captcha.html" in current_url:
        print("\n🚨 CAPTCHA page detected:", current_url)
        print("Please solve the CAPTCHA manually in the browser.")
        input("✅ Press Enter after solving the CAPTCHA...")

    # Case 2: CAPTCHA-related text in page source
    elif any(phrase in page_text for phrase in ["verify you are human", "just a moment", "security check"]):
        print("\n🚨 CAPTCHA phrase detected in page source!")
        print("Please solve the CAPTCHA manually in the browser.")
        input("✅ Press Enter after solving the CAPTCHA...")

# Handle Excel file and row splitting
def get_new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Make", "Model", "Chassis", "chassis-URL", "New_Model", "Engine Model",
        "Engine_capacity", "Prod period", "Body", "Options", "URL",
        "product_code", "product_url", "Product Name", "Part Number",
        "Chassis Details", "Product Period", "Product Status"
    ])
    return wb, ws

output_index = 1
wb, ws = get_new_workbook()
row_counter = 1  # Including header row

df = pd.read_excel(input_path)

for index, row in df.iterrows():
    print(f"Working on {index + 1} / {len(df)}")

    base_info = [
        row["Make"], row["Model"], row["Chassis"], row["chassis-URL"],
        row["New_Model"], row["Engine Model"], row["Engine_capacity"],
        row["Prod period"], row["Body"], row["Options"], row["URL"]
    ]

    try:
        driver.get(row["URL"])
        check_for_captcha()
        time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

        schema_divs = driver.find_elements(By.CSS_SELECTOR, ".epcVariation__schema")
              
        # ✅ Only allow full product_code + name format
        allowed_product_codes = {
            "35-110 - Front wheel brake",
            "35-160 - Rear wheel brake",
            "15-10 - Air cleaner",
            "12-110 - Oil pump & oil filter",
            "11-510 - Engine mounting & support",
            "35-630 - Brake master cylinder",
            "21-10 - Clutch & clutch release",
            "13-40 - Fuel pump",
            "11-90 - Engine overhaul gasket kit",
            "16-220 - Spark plug,cable & coil",
            "14-110 - Water pump",
            "31-10 - Wheel,tire & cover",
            "42-320 - Fuel filler lid & lock"
        }

        product_urls = []
        for div in schema_divs:
            try:
                a_tag = div.find_element(By.CSS_SELECTOR, "a")
                url = a_tag.get_attribute("href")
                code = a_tag.find_element(By.TAG_NAME, "img").get_attribute("alt").strip()
                if code in allowed_product_codes:
                    product_urls.append((code, url))
            except:
                continue

        for code, product_url in product_urls:
            driver.get(product_url)
            check_for_captcha()
            time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

            current_product_name = ""
            group_rows = driver.find_elements(By.CSS_SELECTOR, "table.entriesPncTable tbody tr")

            for tr in group_rows:
                try:
                    tds = tr.find_elements(By.TAG_NAME, "td")
                    if len(tds) == 1 and "groupHeader" in tds[0].get_attribute("class"):
                        current_product_name = tds[0].text.strip()
                    elif len(tds) >= 5:
                        part_number = tds[0].text.strip()
                        chassis_details = tds[1].text.strip()
                        product_period = ""
                        product_status = ""

                        for td in tds:
                            class_attr = td.get_attribute("class")
                            if "entriesTable__period" in class_attr:
                                product_period = td.text.strip()
                            elif "entriesTable__price" in class_attr:
                                product_status = td.text.strip()

                        row_data = base_info + [
                            code, product_url,
                            current_product_name, part_number,
                            chassis_details, product_period, product_status
                        ]
                        ws.append(row_data)
                        row_counter += 1

                        if row_counter >= MAX_ROWS:
                            save_path = f"{output_base}_{output_index}.xlsx"
                            wb.save(save_path)
                            print(f"Saved {save_path} with {row_counter} rows.")
                            output_index += 1
                            wb, ws = get_new_workbook()
                            row_counter = 1

                except Exception as e:
                    print(f"Error parsing row in {product_url}: {e}")
                    continue

    except Exception as e:
        print(f"Error on URL {row['URL']}: {e}")
        continue

# Final Save
final_path = f"{output_base}_{output_index}.xlsx"
wb.save(final_path)
print(f"Final file saved: {final_path}")
driver.quit()
print("✅ Scraping complete!")

